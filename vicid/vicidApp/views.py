from django.shortcuts import render, redirect,get_object_or_404
from django.urls import reverse
from .forms import UserForm, ReportForm
from django.http import HttpResponse
from .serializers import UsersSerializer, CustomersSerializer
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Users, Customer, History
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter 

# Create your views here.

def view_all(request):
    return render(request, 'base.html')

def view_profile(request, user_id):
    user = get_object_or_404(Users, user_id=user_id)
    return render(request, 'profile.html', {'user': user})

def edit_profile(request, user_id):
    user = get_object_or_404(Users, user_id=user_id)

    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES,instance=user)

        if form.is_valid():
            form.save()
            profile_url = reverse('view_profile', kwargs={'user_id': user.user_id})
            return redirect(profile_url)

    else:
        form = UserForm()
        return render(request, 'edit_profile.html', {'form':form, 'user':user})

def success(request):
    return HttpResponse('successfully uploaded')

def download_page(request,user_id):
    user = get_object_or_404(Users, user_id=user_id)
    form = ReportForm()
    return render(request,'download_excel.html',{'form':form,'user':user})

def export_to_excel(request,user_id):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="data_history.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "History"

    headers = ["VICID", "ACCTNO", "NO_IDEN","NAMA_NASABAH","Nama_Field","Field Awal","Field Update","Record_DEL","TGL_UPDT","JM_UPDT","TGL_APPRV_UPDT","JM_APPRV_UPDT","USR_UPDT","APPRV_UPDT","TGL_DELET","JM_DELET","TGL_APPRV_DELET","JM_APPRV_DELET","USR_DELET","APPRV_DELET"]
    ws.append(headers)

    column_widths = [5, 15, 20, 20,15,15,15,15,15,15,15,15,15,15,15,15,15,15,15,15,15,15,15,15] 
    for i, column_width in enumerate(column_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = column_width

    # Add data from the model
    form = ReportForm(request.POST)
    if form.is_valid():
        start_date = form.cleaned_data['StartDate']
        end_date = form.cleaned_data['EndDate']

    print("start date",start_date)

    histories = History.objects.filter(user=user_id,  TGL_UPDT__range=(start_date.strftime("%Y%m%d"), end_date.strftime("%Y%m%d")))
    customer_id = histories.values_list('customer', flat=True).distinct()
    customers = Customer.objects.filter(VICID__in=customer_id)

    for history in histories:
        print("masuk history",histories)
        for customer in customers:
            print("masuk customer", history.customer.VICID)
            if history.customer.VICID == customer.VICID:
                 print("acct no", customer.ACCTNO)
                 ws.append([
                    customer.VICID, 
                    customer.ACCTNO, 
                    customer.NO_IDEN, 
                    customer.NAMA_NASABAH,
                    history.Nama_Field,
                    history.Field_Awal,
                    history.Field_Update,
                    history.Record_DEL,
                    history.TGL_UPDT,
                    history.JM_UPDT,
                    history.TGL_APPRV_UPDT,
                    history.JM_APPRV_UPDT,
                    history.USR_UPDT,
                    history.APPRV_UPDT,
                    history.TGL_DELET,
                    history.JM_DELET,
                    history.TGL_APPRV_DELET,
                    history.JM_APPRV_DELET,
                    history.USR_DELET,
                    history.APPRV_DELET
                    ])

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    wb.save(response)
    return response


# add data user via API
@api_view(['POST'])
def add_user(request):
    serializer = UsersSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def add_customer(request):
    serializer = CustomersSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
